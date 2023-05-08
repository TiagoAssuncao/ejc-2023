import openpyxl
import re

def transform_excel_file(input_file_path, output_file_path):
    """
    Reads an Excel file at the specified input file path, removes numbers and dots from the beginning of the values in the first
    column, removes whitespace and special characters from the values in the second column, and writes the updated values to a new
    Excel file at the specified output file path.

    Args:
        input_file_path (str): The path of the input Excel file.
        output_file_path (str): The path of the output Excel file.

    Returns:
        None
    """

    # Open the input Excel file
    workbook = openpyxl.load_workbook(input_file_path)

    # Select the first worksheet
    sheet = workbook.active

    # Store the values of the second column in a list, cleaned of whitespace and special characters
    column2 = [re.sub(r'\W+', '', str(cell.value).strip()) for cell in sheet['B']]

    # Remove numbers and dots from the values in the first column
    column1_new = []
    for cell in sheet['A']:
        value = str(cell.value).strip()
        if re.match(r'^\d+\.', value):
            value = re.sub(r'^\d+\.', '', value)
        column1_new.append(value)

    # Create a new workbook and sheet with the updated values
    workbook_new = openpyxl.Workbook()
    sheet_new = workbook_new.active

    # Write the updated values to the new sheet
    for i, (value1, value2) in enumerate(zip(column1_new, column2)):
        sheet_new.cell(row=i+1, column=1, value=value1)
        sheet_new.cell(row=i+1, column=2, value=value2)

    # Save the new workbook to a file
    workbook_new.save(output_file_path)


transform_excel_file("file.xlsx", "file_transformation.xlsx")
